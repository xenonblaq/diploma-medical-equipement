import pandas as pd
import os

import warnings
warnings.filterwarnings('ignore', category=FutureWarning)

# OpenPyXL используется только для красивого форматирования итогового Excel-файла:
# - стиль шапки (заливка, жирный шрифт)
# - стиль обычных ячеек (границы, выравнивание)
# - автоширина колонок по содержимому
# - закрепление первой строки
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NamedStyle
from openpyxl.utils import get_column_letter

# Импорты конкретных парсеров.
# Каждый парсер инкапсулирует:
# - parse(): сбор "сырья" (web/API/файлы) + сохранение во временный CSV (почти всегда)
# - decor(): приведение к единому формату колонок, чистки, конвертации валют/единиц, фильтры по годам
from Brazil import Brazil
from Korea import Korea
from Turkey import Turkey
from Japan import Japan
from Kazakhstan import Kazakhstan
from HongKong import HongKong
from Taiwan import Taiwan
from Thailand import Thailand
from USA import USA
from Mexico import Mexico
from EU import EU
from India import India
from Vietnam import Vietnam
from Uzbekistan import Uzbekistan
from Armenia import Armenia
from Tadjikistan import Tadjikistan
from Kyrgyzstan import Kyrgyzstan
from Azerbaijan import Azerbaijan
from CIS import CIS
from Belarus import Belarus
from China import China


class RussianForeignTradeParser_1:
    """Единая обертка-роутер над всеми парсерами.

    Задачи класса:
    1) По названию страны выбрать нужный парсер и запустить его (parse()).
    2) Опционально прогнать все парсеры из списка countries и объединить результат.
    3) Сохранить итоговый DataFrame в Excel с аккуратным форматированием.

    Важно:
    - Конкретные параметры запуска (years, belarus, months, и т.п.) пробрасываются внутрь
      парсеров через self.params (см. __init__ и init_parser).
    - Все парсеры должны в итоге вернуть DataFrame в ЕДИНОМ формате колонок (см. README).
    """

    # Стиль шапки таблицы (первая строка):
    # - белый жирный текст
    # - синяя заливка
    # - тонкие границы со всех сторон
    # - выравнивание по центру
    HEADER_STYLE = NamedStyle(
        name='header',
        font=Font(bold=True, color='FFFFFF'),
        alignment=Alignment(horizontal='center', vertical='center'),
        fill=PatternFill(fill_type='solid', fgColor='4F81BD'),
        border=Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
    )

    # Стиль обычных ячеек:
    # - выравнивание по левому краю
    # - тонкие границы
    CELL_STYLE = NamedStyle(
        name='cell',
        alignment=Alignment(horizontal='left', vertical='center'),
        border=Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
    )

    # Список поддерживаемых стран (ключи для выбора парсера).
    # ВНИМАНИЕ: строки должны совпадать с ветками в init_parser().
    countries = [
        "Brazil", "Korea", "Turkey", "Japan", "Kazakhstan", "Hong Kong",
        "Taiwan", "Thailand", "USA", "Mexico", "EU", "India", "Vietnam",
        "Uzbekistan", "Armenia", "Tadjikistan", "Kyrgyzstan", "Azerbaijan",
        "CIS", "Belarus", "China"
    ]

    def __init__(self, country: str = "", all: bool = False, params: dict = {}) -> None:
        """Инициализация обертки.

        Args:
            country: название страны (как в списке countries).
            all: если True — запускать ВСЕ парсеры из списка countries.
            params: параметры, которые будут переданы в конструктор выбранного парсера
                   (например: {"years": ["2024","2025"]} или {"years": ["2025"], "belarus": True}).
        """
        self.all = all
        self.params = params

        # Если all=False — заранее инициализируем конкретный парсер
        # (чтобы при parse() не заниматься роутингом повторно).
        if not all:
            self.country = country
            self.parser = self.init_parser(country)

    def init_parser(self, country: str):
        """Фабрика парсеров: по названию страны возвращает объект нужного класса парсера.

        Важно:
        - Здесь нет логики парсинга — только маршрутизация.
        - self.params распаковываются как kwargs в конкретный класс парсера.
        """
        if country == "Brazil":
            parser = Brazil(**self.params)
        elif country == "Korea":
            parser = Korea(**self.params)
        elif country == "Turkey":
            parser = Turkey(**self.params)
        elif country == "Japan":
            parser = Japan(**self.params)
        elif country == "Kazakhstan":
            parser = Kazakhstan(**self.params)
        elif country == "Hong Kong":
            parser = HongKong(**self.params)
        elif country == "Taiwan":
            parser = Taiwan(**self.params)
        elif country == "Thailand":
            parser = Thailand(**self.params)
        elif country == "USA":
            parser = USA(**self.params)
        elif country == "Mexico":
            parser = Mexico(**self.params)
        elif country == "EU":
            parser = EU(**self.params)
        elif country == "India":
            parser = India(**self.params)
        elif country == "Vietnam":
            parser = Vietnam(**self.params)
        elif country == "Uzbekistan":
            parser = Uzbekistan(**self.params)
        elif country == "Armenia":
            parser = Armenia(**self.params)
        elif country == "Tadjikistan":
            parser = Tadjikistan(**self.params)
        elif country == "Kyrgyzstan":
            parser = Kyrgyzstan(**self.params)
        elif country == "Azerbaijan":
            parser = Azerbaijan(**self.params)
        elif country == "CIS":
            parser = CIS(**self.params)
        elif country == "Belarus":
            parser = Belarus(**self.params)
        elif country == "China":
            parser = China(**self.params)
        else:
            # Если страна не поддерживается — возвращаем None.
            # parse() ниже выбросит понятную ошибку.
            return None
        return parser

    def parse_all(self) -> pd.DataFrame:
        """Запуск всех парсеров из списка countries и объединение результатов в один DataFrame."""
        dfs = []
        for country in RussianForeignTradeParser_1.countries:
            parser = self.init_parser(country)
            data = parser.parse()
            dfs.append(data)
        return pd.concat(dfs)

    def parse(self) -> pd.DataFrame:
        """Запуск выбранного парсера (или всех парсеров, если all=True)."""
        if self.all:
            return self.parse_all()

        if not self.parser:
            raise AttributeError(
                "Данная страна еще не реализована. "
                "Проверьте, что ваша страна входит в список: "
                + ', '.join(RussianForeignTradeParser_1.countries)
            )

        data = self.parser.parse()
        return data

    def create_excel(self, df: pd.DataFrame, output_filename: str = 'output.xlsx', sheet_name: str = 'Sheet1') -> None:
        """Сохранить DataFrame в Excel с форматированием.

        Что делает:
        - создает папку data/excel_files (если нет)
        - сохраняет df без индекса
        - применяет стили к шапке и ячейкам
        - подбирает ширину колонок по максимальной длине значений
        - фиксирует шапку (freeze panes)
        """
        os.makedirs('data/excel_files', exist_ok=True)
        filepath = os.path.join('data/excel_files', output_filename)

        writer = pd.ExcelWriter(filepath, engine='openpyxl')
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        wb = writer.book
        ws = writer.sheets[sheet_name]

        # Регистрируем стили (если еще не зарегистрированы в книге)
        if 'header' not in wb.named_styles:
            wb.add_named_style(self.HEADER_STYLE)
        if 'cell' not in wb.named_styles:
            wb.add_named_style(self.CELL_STYLE)

        # Шапка (первая строка)
        for cell in ws[1]:
            cell.style = 'header'

        # Остальные ячейки (со 2-й строки)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.style = 'cell'

        # Автоширина колонок: берем максимум длины текста в колонке и длину заголовка
        for idx, col in enumerate(df.columns, 1):
            max_length = max(df[col].astype(str).map(len).max(), len(col))
            ws.column_dimensions[get_column_letter(idx)].width = max_length + 2

        # Закрепляем шапку
        ws.freeze_panes = ws['A2']

        writer.close()
        print(f"Финальная таблица '{output_filename}' находится в папке 'data/excel_files'.")
